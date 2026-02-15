import re
import pandas as pd
from typing import Dict, Any, List
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from app.brands.base import BaseBrandProcessor


class AlArabiProcessor(BaseBrandProcessor):
    brand_name = "Al Arabi"
    brand_name_ar = "أقمشة العربي"

    @staticmethod
    def _extract_valid_sku_qty(product_list):
        result = []
        for item in product_list:
            sku_match = re.search(r'SKU:\s*([\w\d_]+)', item)
            qty_match = re.search(r'Qty:\s*(\d+)', item)
            if sku_match and qty_match:
                sku = sku_match.group(1)
                qty = int(qty_match.group(1))
                result.append([sku, qty])
        return result

    @staticmethod
    def _calculate_boxes(product, white, cremy):
        total_units = white + cremy
        if product == 'N':
            return round(total_units / 2, 2)
        return round(total_units / 3, 2)

    def _count_skus(self, sku_qty_pairs: list) -> defaultdict:
        """Core counting logic shared by file and DB modes."""
        totals = defaultdict(lambda: {"White": 0, "Cremy": 0})
        for sku_code, qty in sku_qty_pairs:
            if not sku_code:
                continue
            match = re.match(r"([A-Z]+)_(.*)", str(sku_code).upper())
            if not match:
                continue
            product, components = match.groups()
            parts = re.findall(r"(\d+)([WC])", components)
            for count, color in parts:
                if color == "W":
                    totals[product]["White"] += int(count) * qty
                elif color == "C":
                    totals[product]["Cremy"] += int(count) * qty
        return totals

    def _build_result(self, totals: defaultdict, total_orders: int) -> Dict[str, Any]:
        rows = []
        total_white = total_cremy = total_boxes = 0
        for product in sorted(totals.keys()):
            w = totals[product]["White"]
            c = totals[product]["Cremy"]
            taka_w = round(w / 6, 2)
            taka_c = round(c / 6, 2)
            boxes = self._calculate_boxes(product, w, c)
            total_white += w
            total_cremy += c
            total_boxes += boxes
            rows.append({"المنتج": product, "أبيض (White)": w, "كريمي (Cremy)": c, "طاقة الأبيض": taka_w, "طاقة الكريمي": taka_c, "الكراتين": boxes})

        rows.append({"المنتج": "المجموع", "أبيض (White)": total_white, "كريمي (Cremy)": total_cremy, "طاقة الأبيض": round(total_white / 6, 2), "طاقة الكريمي": round(total_cremy / 6, 2), "الكراتين": round(total_boxes, 2)})

        return {
            "tables": [{"title": "ملخص المنتجات - أقمشة العربي", "columns": ["المنتج", "أبيض (White)", "كريمي (Cremy)", "طاقة الأبيض", "طاقة الكريمي", "الكراتين"], "rows": rows}],
            "summary": {"total_orders": total_orders, "total_white": total_white, "total_cremy": total_cremy}
        }

    def compute_tables(self, df: pd.DataFrame) -> Dict[str, Any]:
        sku_col = "اسماء المنتجات مع SKU"
        if sku_col not in df.columns:
            for col in df.columns:
                if "sku" in col.lower() or "منتج" in col:
                    sku_col = col
                    break
            else:
                raise KeyError("الملف لا يحتوي على عمود المنتجات/SKU المطلوب")

        df[sku_col] = df[sku_col].astype(str).str.split(',')
        df['cleaned_SKU1'] = df[sku_col].apply(self._extract_valid_sku_qty)

        all_pairs = []
        for row in df['cleaned_SKU1']:
            all_pairs.extend(row)

        totals = self._count_skus(all_pairs)
        return self._build_result(totals, len(df))

    def compute_from_sku_list(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        pairs = [(item["sku"], item["quantity"]) for item in items]
        totals = self._count_skus(pairs)
        return self._build_result(totals, 0)

    def export_excel(self, tables: Dict[str, Any], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Al Arabi Report"
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        total_fill = PatternFill(start_color="DEB887", end_color="DEB887", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        current_row = 1
        for table in tables.get("tables", []):
            ws.cell(row=current_row, column=1, value=table["title"]).font = Font(bold=True, size=13)
            current_row += 1

            for col_idx, col_name in enumerate(table["columns"], 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            for row_data in table["rows"]:
                is_total = row_data.get("المنتج") == "المجموع"
                for col_idx, col_name in enumerate(table["columns"], 1):
                    val = row_data.get(col_name, "")
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if is_total:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)
                current_row += 1

            current_row += 2

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        return output_path
