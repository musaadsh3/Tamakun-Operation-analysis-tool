import re
import pandas as pd
from typing import Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from app.brands.base import BaseBrandProcessor


class ShabahProcessor(BaseBrandProcessor):
    brand_name = "Shabah"
    brand_name_ar = "شبة"

    @staticmethod
    def _extract_codes(sku: str):
        sku = str(sku)
        htb = fhm = fire_starter = hidroleck = bbq = inc = None

        m_htb = re.search(r'HTB(?:W?\d+Q?\d*T)?(\d+)', sku)
        if m_htb:
            htb = "HTB" + m_htb.group(1)

        m_fhm = re.search(r'FHM(\d+)', sku)
        if m_fhm:
            fhm = "FHM" + m_fhm.group(1)

        if re.search(r'FS_bag_Q20', sku):
            fire_starter = 'مشعل النار'

        if re.search(r'Car_Jack_5_ton', sku):
            hidroleck = 'طقم هايدروليك 3 في 1'

        if re.search(r'FHM-BBQ', sku):
            bbq = 'فحم الشواء'

        if re.search(r'FHM-INC', sku):
            inc = 'فحم البخور'

        return htb, fhm, fire_starter, hidroleck, bbq, inc

    @staticmethod
    def _parse_names_column(cell):
        if cell is None or (isinstance(cell, float) and pd.isna(cell)):
            return []
        s = str(cell).replace("\n", " ").replace("\r", " ")
        matches = re.findall(
            r"\(SKU:\s*([A-Za-z0-9_]+)\).*?\(Qty:\s*(\d+)\)", s
        )
        return [(sku, int(qty)) for sku, qty in matches]

    def _count_skus(self, sku_qty_pairs: List[Tuple[str, int]]) -> Dict[str, Any]:
        """Core counting logic shared by file and DB modes."""
        totals = {}
        def add_qty(key, qty):
            totals[key] = totals.get(key, 0) + qty

        for sku, qty in sku_qty_pairs:
            if not sku:
                continue
            htb, fhm, fire_starter, hidroleck, bbq, inc = self._extract_codes(sku)
            if htb:
                add_qty(htb, qty)
            if fhm:
                add_qty(fhm, qty)
            if fire_starter:
                add_qty(fire_starter, qty)
            if hidroleck:
                add_qty(hidroleck, qty)
            if bbq:
                add_qty(bbq, qty)
            if inc:
                add_qty(inc, qty)

        return totals

    def _build_result(self, totals: dict, total_orders: int) -> Dict[str, Any]:
        rows = sorted(
            [{"SKU": k, "الكمية": v} for k, v in totals.items()],
            key=lambda x: x["SKU"]
        )
        total_qty = sum(r["الكمية"] for r in rows)
        rows.append({"SKU": "المجموع", "الكمية": total_qty})
        return {
            "tables": [{"title": "ملخص المنتجات - شبة", "columns": ["SKU", "الكمية"], "rows": rows}],
            "summary": {"total_orders": total_orders, "total_products": total_qty}
        }

    def compute_tables(self, df: pd.DataFrame) -> Dict[str, Any]:
        sku_col = "اسماء المنتجات مع SKU"
        if sku_col not in df.columns:
            for col in df.columns:
                if "sku" in col.lower() or "منتج" in col:
                    sku_col = col
                    break
            else:
                raise KeyError("الملف لا يحتوي على عمود المنتجات/SKU المطلوب")

        all_pairs = []
        for cell in df[sku_col]:
            items = self._parse_names_column(cell)
            all_pairs.extend(items)

        totals = self._count_skus(all_pairs)
        return self._build_result(totals, len(df))

    def compute_from_sku_list(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        pairs = [(item["sku"], item["quantity"]) for item in items]
        totals = self._count_skus(pairs)
        return self._build_result(totals, 0)

    def export_excel(self, tables: Dict[str, Any], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Shabah Report"
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        total_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        current_row = 1
        for table in tables.get("tables", []):
            ws.cell(row=current_row, column=1, value=table["title"]).font = Font(bold=True, size=13)
            current_row += 1

            for col_idx, col_name in enumerate(table["columns"], 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            for row_data in table["rows"]:
                is_total = row_data.get("SKU") == "المجموع"
                for col_idx, col_name in enumerate(table["columns"], 1):
                    val = row_data.get(col_name, "")
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if is_total:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)
                current_row += 1

            current_row += 2

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        return output_path
