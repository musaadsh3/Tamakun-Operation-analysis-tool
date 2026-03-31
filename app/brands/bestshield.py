import re
import json
import pandas as pd
from typing import Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from app.brands.base import BaseBrandProcessor


class BestShieldProcessor(BaseBrandProcessor):
    brand_name = "Best Shield"
    brand_name_ar = "بست شيلد"

    NUMBERS = ['70', '50', '35', '20', '05']

    # ── Helpers ─────────────────────────────────────────────

    @staticmethod
    def _split_sku(sku: str) -> List[str]:
        """Split a SKU string into parts, preserving ppf_ prefixes."""
        cleaned = str(sku).replace("'", "").replace('"', "").strip(',').strip()
        return re.split(r'(?<!ppf)_', cleaned)

    # ── Excel file parsers ───────────────────────────────────

    @staticmethod
    def _extract_skus_from_item(item) -> List[Tuple[str, int]]:
        """
        Recursively extract (sku, qty) pairs from a single item structure.
        item: [name, qty, sku] or [name, qty, sku, [sub_items...]]
        For bundles (4th element is a list), we IGNORE the bundle SKU
        and only process sub-items.
        """
        result = []
        if not isinstance(item, list) or len(item) < 3:
            return result
        sku = item[2]
        qty = item[1]
        if len(item) >= 4 and isinstance(item[3], list):
            for sub in item[3]:
                result.extend(BestShieldProcessor._extract_skus_from_item(sub))
        else:
            result.append((sku, qty))
        return result

    @staticmethod
    def _parse_salla_cell(cell_value) -> List[Tuple[List[str], int]]:
        """
        Parse JSON cell from skus_json column.
        Returns [(parts_list, qty), ...] where parts_list is already split.
        """
        if not isinstance(cell_value, list):
            try:
                if pd.isna(cell_value):
                    return []
            except (TypeError, ValueError):
                pass
        if isinstance(cell_value, list):
            data = cell_value
        else:
            try:
                data = json.loads(str(cell_value))
            except Exception:
                return []

        all_pairs = []
        for item in data:
            if isinstance(item, list):
                all_pairs.extend(BestShieldProcessor._extract_skus_from_item(item))

        result = []
        for sku, qty in all_pairs:
            if not isinstance(sku, str):
                continue
            try:
                qty_int = int(qty)
            except Exception:
                continue
            parts = BestShieldProcessor._split_sku(sku)
            result.append((parts, qty_int))
        return result

    @staticmethod
    def _parse_names_column(cell) -> List[Tuple[List[str], int]]:
        """
        Parse '(SKU: X)(Qty: Y)' format.
        Returns [(parts_list, qty), ...] where parts_list is already split.
        """
        if cell is None or (isinstance(cell, float) and pd.isna(cell)):
            return []
        s = str(cell).replace("\n", " ").replace("\r", " ")
        matches = re.findall(r"\(SKU:\s*([A-Za-z0-9_\-]+)\).*?\(Qty:\s*(\d+)\)", s)
        result = []
        for sku, qty in matches:
            parts = BestShieldProcessor._split_sku(sku)
            result.append((parts, int(qty)))
        return result

    # ── Core SKU counting (shared by file & DB modes) ────────

    def _count_skus(self, parsed_items: List[Tuple[List[str], int]]) -> Dict[str, Any]:
        """
        Process parsed SKU items and count all products.
        parsed_items: [(parts_list, qty), ...] where parts_list is a list of strings.
        """
        numbers = self.NUMBERS  # ['70', '50', '35', '20', '05']

        # ── Normal counters ──
        B = {f'B{n}': 0 for n in numbers}
        S = {f'S{n}': 0 for n in numbers}
        ppf = {"ppf_q": 0, "ppf_e": 0, "ppf_f": 0, "ppf_p": 0}
        H_dict = {
            f'H{h}': {f'T{t}': 0 for t in ['70', '50', '35', '20', '5']}
            for h in [10, 5, 3, 1]
        }

        protect_shield = 0
        CARD = 0
        SPRO3 = 0
        mic3 = 0
        CLNo3 = 0
        TINT_BOX_FORGET = 0
        brake_front = 0
        brake_rear = 0
        dashCam = 0

        # ── B2B counters ──
        B_b2b = {f'B{n}': 0 for n in numbers}
        S_b2b = {f'S{n}': 0 for n in numbers}
        ppf_b2b = {"ppf_q": 0, "ppf_e": 0, "ppf_f": 0, "ppf_p": 0}
        H_dict_b2b = {
            f'H{h}': {f'T{t}': 0 for t in ['70', '50', '35', '20', '5']}
            for h in [10, 5, 3, 1]
        }

        protect_shield_b2b = 0
        CARD_b2b = 0
        SPRO3_b2b = 0
        mic3_b2b = 0
        CLNo3_b2b = 0
        TINT_BOX_FORGET_b2b = 0
        brake_front_b2b = 0
        brake_rear_b2b = 0

        # ── Old SKU mappings ──
        B_old = {"B00": "B70", "B01": "B50", "B02": "B35", "B03": "B20"}
        S_old = {"S00": "S70", "S01": "S50", "S02": "S35", "S03": "S20"}
        old_sku_mapping = {"0": "70", "1": "50", "2": "35", "3": "20"}

        # ── Tracking counters ──
        pdz = 0
        I_count = 0
        II_count = 0
        BR = 0
        SR = 0
        HH = 0

        pdz_b2b = 0
        I_b2b = 0
        II_b2b = 0
        BR_b2b = 0
        SR_b2b = 0
        HH_b2b = 0

        for j, qty in parsed_items:
            if not isinstance(j, list) or not j:
                continue

            # B2B detection (case-insensitive)
            is_b2b = False
            if str(j[0]).upper() == 'B2B':
                is_b2b = True
                j = j[1:]

            if not j:
                continue

            # ── Combo SKUs: I/II/PDZ + B + S + 2S + B ──
            if len(j) >= 5 and j[0] in ["I", "II", "PDZ"]:
                first, second, third, fourth, fifth = j[:5]
                if (
                    second.startswith("B") and second[1:] in numbers
                    and third.startswith("S") and third[1:] in numbers
                    and fourth.startswith("2S") and fourth[2:] in numbers
                    and fifth.startswith("B") and fifth[1:] in numbers
                ):
                    if is_b2b:
                        if first == 'PDZ':
                            pdz_b2b += qty
                            TINT_BOX_FORGET_b2b += qty
                            protect_shield_b2b += qty
                        elif first == 'II':
                            II_b2b += qty
                        elif first == 'I':
                            I_b2b += qty
                        mult = 2 if first == 'II' else 1
                        B_b2b[second] += qty * mult
                        B_b2b[fifth] += qty * mult
                        conv_fourth = "S" + fourth[2:]
                        S_b2b[third] += qty * mult
                        S_b2b[conv_fourth] += qty * 2 * mult
                    else:
                        if first == 'PDZ':
                            pdz += qty
                            TINT_BOX_FORGET += qty
                            protect_shield += qty
                        elif first == 'II':
                            II_count += qty
                        elif first == 'I':
                            I_count += qty
                        mult = 2 if first == 'II' else 1
                        B[second] += qty * mult
                        B[fifth] += qty * mult
                        conv_fourth = "S" + fourth[2:]
                        S[third] += qty * mult
                        S[conv_fourth] += qty * 2 * mult
                    continue

            # ── Old combo SKUs (G/S type) ──
            if len(j) == 5 and j[0] in ["G", "S"]:
                first, second, third, fourth, fifth = j[:5]
                if (
                    second.startswith("AA") and second[-4:] in ["2000", "1000"]
                    and third.startswith("JA") and third[-4:] in ["2000", "1000"]
                    and fourth.startswith("JK") and fourth[-4:] in ["2000", "1000"]
                    and fifth.startswith("KK") and fifth[-4:] in ["2000", "1000"]
                ):
                    second_key = f'B{old_sku_mapping.get(second[2], "INVALID")}'
                    third_key = f'S{old_sku_mapping.get(third[2], "INVALID")}'
                    fourth_key = f'B{old_sku_mapping.get(fourth[2], "INVALID")}'
                    fifth_key = f'B{old_sku_mapping.get(fifth[2], "INVALID")}'

                    I_count += qty
                    if second_key in B:
                        B[second_key] += qty
                    if third_key in S:
                        S[third_key] += qty
                    if fourth_key in B:
                        B[fourth_key] += qty
                    if fifth_key in B:
                        B[fifth_key] += qty
                    continue

            # ── DZ shield-only SKU (e.g., DZ_H150_W100) ──
            if j[0] == "DZ" and len(j) == 3:
                if is_b2b:
                    protect_shield_b2b += qty
                else:
                    protect_shield += qty
                continue

            # ── Single item SKUs (len == 1) ──
            if len(j) == 1:
                item = j[0]
                if is_b2b:
                    if item in B_b2b:
                        B_b2b[item] += qty
                        BR_b2b += qty
                    elif item in B_old:
                        B_b2b[B_old[item]] += qty
                    elif item in S_b2b:
                        S_b2b[item] += qty
                        SR_b2b += qty
                    elif item in S_old:
                        S_b2b[S_old[item]] += qty
                    elif re.match(r'^(H\d+)W1T(\d+)$', item):
                        m = re.match(r'^(H\d+)W1T(\d+)$', item)
                        if m:
                            HH_b2b += qty
                            h_val, t_val = m.groups()
                            t_key = f'T{t_val}'
                            if h_val in H_dict_b2b and t_key in H_dict_b2b[h_val]:
                                H_dict_b2b[h_val][t_key] += qty
                    elif item in ["SPR03", "SPR3"]:
                        SPRO3_b2b += qty
                    elif item == "CARD":
                        CARD_b2b += qty
                    elif item == "mic3":
                        mic3_b2b += qty
                    elif item == "CLNo3":
                        CLNo3_b2b += qty
                    elif item == "ppf_q":
                        ppf_b2b["ppf_q"] += qty
                    elif item == "ppf_p":
                        ppf_b2b["ppf_p"] += qty
                    elif item == "ppf_e":
                        ppf_b2b["ppf_e"] += qty
                    elif item == "ppf_f":
                        ppf_b2b["ppf_f"] += qty
                    elif re.match(r'^D\d+-(F|R)-\d+$', item):
                        side = re.findall(r'^D\d+-(F|R)-\d+$', item)[0]
                        if side == "F":
                            brake_front_b2b += qty
                        else:
                            brake_rear_b2b += qty
                else:
                    if item in B:
                        B[item] += qty
                        BR += qty
                    elif item in B_old:
                        B[B_old[item]] += qty
                    elif item in S:
                        S[item] += qty
                        SR += qty
                    elif item in S_old:
                        S[S_old[item]] += qty
                    elif re.match(r'^(H\d+)W1T(\d+)$', item):
                        m = re.match(r'^(H\d+)W1T(\d+)$', item)
                        if m:
                            HH += qty
                            h_val, t_val = m.groups()
                            t_key = f'T{t_val}'
                            if h_val in H_dict and t_key in H_dict[h_val]:
                                H_dict[h_val][t_key] += qty
                    elif re.match(r'^D\d+-(F|R)-\d+$', item):
                        side = re.findall(r'^D\d+-(F|R)-\d+$', item)[0]
                        if side == "F":
                            brake_front += qty
                        else:
                            brake_rear += qty
                    elif item in ["SPR03", "SPR3"]:
                        SPRO3 += qty
                    elif item == "CARD":
                        CARD += qty
                    elif item == "mic3":
                        mic3 += qty
                    elif item == "CLNo3":
                        CLNo3 += qty
                    elif item == "ppf_q":
                        ppf["ppf_q"] += qty
                    elif item == "ppf_p":
                        ppf["ppf_p"] += qty
                    elif item == "ppf_e":
                        ppf["ppf_e"] += qty
                    elif item == "ppf_f":
                        ppf["ppf_f"] += qty
                    elif item == "DC-4K-W":
                        dashCam += qty
                continue

        # Rename T5 -> T05 for consistency in output
        for h in H_dict:
            H_dict[h]['T05'] = H_dict[h].pop('T5')
        for h in H_dict_b2b:
            H_dict_b2b[h]['T05'] = H_dict_b2b[h].pop('T5')

        return {
            "B": B, "S": S, "ppf": ppf, "H_dict": H_dict,
            "protect_shield": protect_shield, "CARD": CARD,
            "SPRO3": SPRO3, "mic3": mic3, "CLNo3": CLNo3,
            "TINT_BOX_FORGET": TINT_BOX_FORGET,
            "brake_front": brake_front, "brake_rear": brake_rear,
            "dashCam": dashCam,
            "pdz": pdz, "I": I_count, "II": II_count,
            "BR": BR, "SR": SR, "HH": HH,
            # B2B
            "B_b2b": B_b2b, "S_b2b": S_b2b, "ppf_b2b": ppf_b2b,
            "H_dict_b2b": H_dict_b2b,
            "protect_shield_b2b": protect_shield_b2b, "CARD_b2b": CARD_b2b,
            "SPRO3_b2b": SPRO3_b2b, "mic3_b2b": mic3_b2b,
            "CLNo3_b2b": CLNo3_b2b,
            "TINT_BOX_FORGET_b2b": TINT_BOX_FORGET_b2b,
            "brake_front_b2b": brake_front_b2b, "brake_rear_b2b": brake_rear_b2b,
            "pdz_b2b": pdz_b2b, "I_b2b": I_b2b, "II_b2b": II_b2b,
            "BR_b2b": BR_b2b, "SR_b2b": SR_b2b, "HH_b2b": HH_b2b,
        }

    # ── Build all result tables ─────────────────────────────

    @staticmethod
    def _safe_pct(value, total):
        """Calculate percentage safely, returning 0 if total is 0."""
        if total == 0:
            return 0
        return round(value / total * 100, 2)

    def _build_result_tables(self, c: Dict[str, Any], total_orders: int) -> Dict[str, Any]:
        B, S, ppf, H_dict = c["B"], c["S"], c["ppf"], c["H_dict"]
        B_b2b, S_b2b, ppf_b2b, H_dict_b2b = c["B_b2b"], c["S_b2b"], c["ppf_b2b"], c["H_dict_b2b"]

        sizes = ["70", "50", "35", "20", "05"]
        tables = []

        # ════════════════════════════════════════════════════════
        # 1. Normal B vs S (with percentages)
        # ════════════════════════════════════════════════════════
        total_b = sum(B.values())
        total_s = sum(S.values())
        grand_total = total_b + total_s

        bs_rows = [
            {**{"النوع": "B"}, **{sz: B[f'B{sz}'] for sz in sizes}, "المجموع": total_b},
            {**{"النوع": "S"}, **{sz: S[f'S{sz}'] for sz in sizes}, "المجموع": total_s},
            {**{"النوع": "B %"}, **{sz: self._safe_pct(B[f'B{sz}'], grand_total) for sz in sizes},
             "المجموع": self._safe_pct(total_b, grand_total)},
            {**{"النوع": "S %"}, **{sz: self._safe_pct(S[f'S{sz}'], grand_total) for sz in sizes},
             "المجموع": self._safe_pct(total_s, grand_total)},
        ]
        tables.append({
            "title": "تظليل (B vs S)",
            "columns": ["النوع"] + sizes + ["المجموع"],
            "rows": bs_rows,
        })

        # ════════════════════════════════════════════════════════
        # 2. B2B B vs S (with percentages)
        # ════════════════════════════════════════════════════════
        total_b_b2b = sum(B_b2b.values())
        total_s_b2b = sum(S_b2b.values())
        grand_total_b2b = total_b_b2b + total_s_b2b

        bs_b2b_rows = [
            {**{"النوع": "B B2B"}, **{sz: B_b2b[f'B{sz}'] for sz in sizes}, "المجموع": total_b_b2b},
            {**{"النوع": "S B2B"}, **{sz: S_b2b[f'S{sz}'] for sz in sizes}, "المجموع": total_s_b2b},
            {**{"النوع": "B % B2B"}, **{sz: self._safe_pct(B_b2b[f'B{sz}'], grand_total_b2b) for sz in sizes},
             "المجموع": self._safe_pct(total_b_b2b, grand_total_b2b)},
            {**{"النوع": "S % B2B"}, **{sz: self._safe_pct(S_b2b[f'S{sz}'], grand_total_b2b) for sz in sizes},
             "المجموع": self._safe_pct(total_s_b2b, grand_total_b2b)},
        ]
        tables.append({
            "title": "تظليل B2B",
            "columns": ["النوع"] + sizes + ["المجموع"],
            "rows": bs_b2b_rows,
        })

        # ════════════════════════════════════════════════════════
        # 3. H# table (hardness matrix)
        # ════════════════════════════════════════════════════════
        t_cols = ["T70", "T50", "T35", "T20", "T05"]
        h_rows = []
        for h_key in ['H10', 'H5', 'H3', 'H1']:
            row = {"الصلابة": h_key}
            for t_key in t_cols:
                row[t_key] = H_dict[h_key][t_key]
            row["المجموع"] = sum(H_dict[h_key].values())
            h_rows.append(row)
        tables.append({
            "title": "الصلابة (H)",
            "columns": ["الصلابة"] + t_cols + ["المجموع"],
            "rows": h_rows,
        })

        # ════════════════════════════════════════════════════════
        # 4. Combined result: tint meters + rolls (Normal)
        # ════════════════════════════════════════════════════════
        combined_result = {}
        for sz in sizes:
            vlt = int(sz)
            combined_result[vlt] = (B[f'B{sz}'] * 2) + S[f'S{sz}']

        roll_pkg = {k: round(v / 45, 1) for k, v in combined_result.items()}
        total_tint_rolls = sum(roll_pkg.values())

        tint_roll_rows = []
        for sz in sizes:
            vlt = int(sz)
            tint_roll_rows.append({
                "VLT": vlt,
                "عدد الامتار للتظليل": combined_result[vlt],
                "عدد رولات التظليل": roll_pkg[vlt],
                "النسبة المئوية": self._safe_pct(roll_pkg[vlt], total_tint_rolls),
            })
        # Add total row
        tint_roll_rows.append({
            "VLT": "المجموع",
            "عدد الامتار للتظليل": sum(combined_result.values()),
            "عدد رولات التظليل": round(total_tint_rolls, 1),
            "النسبة المئوية": 100.0,
        })
        tables.append({
            "title": "رولات التظليل",
            "columns": ["VLT", "عدد الامتار للتظليل", "عدد رولات التظليل", "النسبة المئوية"],
            "rows": tint_roll_rows,
        })

        # ════════════════════════════════════════════════════════
        # 5. Combined result: tint meters + rolls (B2B)
        # ════════════════════════════════════════════════════════
        combined_result_b2b = {}
        for sz in sizes:
            vlt = int(sz)
            combined_result_b2b[vlt] = (B_b2b[f'B{sz}'] * 2) + S_b2b[f'S{sz}']

        roll_pkg_b2b = {k: round(v / 45, 1) for k, v in combined_result_b2b.items()}
        total_tint_rolls_b2b = sum(roll_pkg_b2b.values())

        tint_roll_b2b_rows = []
        for sz in sizes:
            vlt = int(sz)
            tint_roll_b2b_rows.append({
                "VLT": vlt,
                "عدد الامتار B2B": combined_result_b2b[vlt],
                "عدد الرولات B2B": roll_pkg_b2b[vlt],
                "النسبة المئوية": self._safe_pct(roll_pkg_b2b[vlt], total_tint_rolls_b2b),
            })
        # Add total row
        tint_roll_b2b_rows.append({
            "VLT": "المجموع",
            "عدد الامتار B2B": sum(combined_result_b2b.values()),
            "عدد الرولات B2B": round(total_tint_rolls_b2b, 1),
            "النسبة المئوية": 100.0,
        })
        tables.append({
            "title": "رولات التظليل B2B",
            "columns": ["VLT", "عدد الامتار B2B", "عدد الرولات B2B", "النسبة المئوية"],
            "rows": tint_roll_b2b_rows,
        })

        # ════════════════════════════════════════════════════════
        # 6. Building (h_calculation) meters + rolls
        # ════════════════════════════════════════════════════════
        h_calculation = {}
        for t_key in ["T70", "T50", "T35", "T20", "T05"]:
            vlt = int(t_key[1:])
            h_calculation[vlt] = sum(
                H_dict[h][t_key] * int(h[1:]) for h in H_dict
            )

        roll_building = {k: round(v / 45, 1) for k, v in h_calculation.items()}
        total_building_rolls = sum(roll_building.values())

        building_rows = []
        for sz in sizes:
            vlt = int(sz)
            building_rows.append({
                "VLT": vlt,
                "عدد الامتار للمباني": h_calculation[vlt],
                "عدد رولات المباني": roll_building[vlt],
                "النسبة المئوية": self._safe_pct(roll_building[vlt], total_building_rolls),
            })
        tables.append({
            "title": "رولات المباني",
            "columns": ["VLT", "عدد الامتار للمباني", "عدد رولات المباني", "النسبة المئوية"],
            "rows": building_rows,
        })

        # ════════════════════════════════════════════════════════
        # 7. Roll total (tint + building)
        # ════════════════════════════════════════════════════════
        roll_total = {int(sz): roll_pkg[int(sz)] + roll_building[int(sz)] for sz in sizes}
        total_all_rolls = sum(roll_total.values())

        roll_total_rows = []
        for sz in sizes:
            vlt = int(sz)
            pct = self._safe_pct(roll_total[vlt], total_all_rolls)
            roll_total_rows.append({
                "VLT": vlt,
                "مجموع الرولات (مباني + تظليل)": roll_total[vlt],
                "النسبة المئوية": f"{pct}%",
            })
        tables.append({
            "title": "مجموع الرولات",
            "columns": ["VLT", "مجموع الرولات (مباني + تظليل)", "النسبة المئوية"],
            "rows": roll_total_rows,
        })

        # ════════════════════════════════════════════════════════
        # 8. PPF (Normal)
        # ════════════════════════════════════════════════════════
        ppf_data = [
            ("ppf_e", ppf['ppf_e'], 4),
            ("ppf_q", ppf['ppf_q'], 8),
            ("ppf_f", ppf['ppf_f'], 21),
            ("ppf_p", ppf['ppf_p'], 1),
        ]
        ppf_rows = []
        for name, cartons, meter_mult in ppf_data:
            ppf_rows.append({
                "المنتج": name,
                "الكراتين": cartons,
                "الامتار": meter_mult * cartons,
                "الرولات": round((meter_mult * cartons) / 15, 2),
            })
        tables.append({
            "title": "PPF",
            "columns": ["المنتج", "الكراتين", "الامتار", "الرولات"],
            "rows": ppf_rows,
        })

        # ════════════════════════════════════════════════════════
        # 9. PPF B2B
        # ════════════════════════════════════════════════════════
        ppf_b2b_data = [
            ("ppf_e", ppf_b2b['ppf_e'], 4),
            ("ppf_q", ppf_b2b['ppf_q'], 8),
            ("ppf_f", ppf_b2b['ppf_f'], 21),
            ("ppf_p", ppf_b2b['ppf_p'], 1),
        ]
        ppf_b2b_rows = []
        for name, cartons, meter_mult in ppf_b2b_data:
            ppf_b2b_rows.append({
                "المنتج": name,
                "الكراتين B2B": cartons,
                "الامتار B2B": meter_mult * cartons,
                "الرولات B2B": round((meter_mult * cartons) / 15, 2),
            })
        tables.append({
            "title": "PPF B2B",
            "columns": ["المنتج", "الكراتين B2B", "الامتار B2B", "الرولات B2B"],
            "rows": ppf_b2b_rows,
        })

        # ════════════════════════════════════════════════════════
        # 10. Big core / Small core (Normal)
        # ════════════════════════════════════════════════════════
        total_B_val = sum(B.values())
        total_T_val = sum(val for h in H_dict for val in H_dict[h].values())
        big_core_total = total_B_val + total_T_val
        small_core_total = sum(S.values())

        tables.append({
            "title": "الكور",
            "columns": ["Big Core Total", "Small Core Total"],
            "rows": [{"Big Core Total": big_core_total, "Small Core Total": small_core_total}],
        })

        # ════════════════════════════════════════════════════════
        # 11. Big core / Small core (B2B)
        # ════════════════════════════════════════════════════════
        total_B_b2b_val = sum(B_b2b.values())
        total_T_b2b_val = sum(val for h in H_dict_b2b for val in H_dict_b2b[h].values())
        big_core_total_b2b = total_B_b2b_val + total_T_b2b_val
        small_core_total_b2b = sum(S_b2b.values())

        tables.append({
            "title": "الكور B2B",
            "columns": ["Big Core Total B2B", "Small Core Total B2B"],
            "rows": [{"Big Core Total B2B": big_core_total_b2b, "Small Core Total B2B": small_core_total_b2b}],
        })

        # ════════════════════════════════════════════════════════
        # 12. الدرع الزجاجي (Normal)
        # ════════════════════════════════════════════════════════
        shield_num = c["protect_shield"]
        shield_meters = round(shield_num * 1.8, 2)
        shield_rolls = round(shield_meters / 30, 2)

        tables.append({
            "title": "الدرع الزجاجي",
            "columns": ["الدرع الزجاجي", "الدرع بالمتر", "عدد رولات الدرع"],
            "rows": [{"الدرع الزجاجي": shield_num, "الدرع بالمتر": shield_meters, "عدد رولات الدرع": shield_rolls}],
        })

        # ════════════════════════════════════════════════════════
        # 13. الدرع الزجاجي B2B
        # ════════════════════════════════════════════════════════
        shield_num_b2b = c["protect_shield_b2b"]
        shield_meters_b2b = round(shield_num_b2b * 1.8, 2)
        shield_rolls_b2b = round(shield_meters_b2b / 30, 2)

        tables.append({
            "title": "الدرع الزجاجي B2B",
            "columns": ["الدرع الزجاجي B2B", "الدرع بالمتر B2B", "عدد رولات الدرع B2B"],
            "rows": [{
                "الدرع الزجاجي B2B": shield_num_b2b,
                "الدرع بالمتر B2B": shield_meters_b2b,
                "عدد رولات الدرع B2B": shield_rolls_b2b,
            }],
        })

        # ════════════════════════════════════════════════════════
        # 14. عدد البكجات (Normal)
        # ════════════════════════════════════════════════════════
        pkg_rows = [
            {"اسم المنتج": "تظليل", "عدد البكجات": c["TINT_BOX_FORGET"] + c["I"] + c["II"] * 2 + c["BR"] + c["SR"]},
            {"اسم المنتج": "مباني", "عدد البكجات": c["HH"]},
            {"اسم المنتج": "درع", "عدد البكجات": c["protect_shield"]},
            {"اسم المنتج": "ppf_q&e&p", "عدد البكجات": ppf['ppf_e'] + ppf['ppf_q'] + ppf['ppf_p']},
            {"اسم المنتج": "ppf_f", "عدد البكجات": ppf['ppf_f']},
        ]
        tables.append({
            "title": "عدد البكجات",
            "columns": ["اسم المنتج", "عدد البكجات"],
            "rows": pkg_rows,
        })

        # ════════════════════════════════════════════════════════
        # 15. عدد البكجات B2B
        # ════════════════════════════════════════════════════════
        pkg_b2b_rows = [
            {"اسم المنتج": "تظليل B2B", "عدد البكجات": c["TINT_BOX_FORGET_b2b"] + c["I_b2b"] + c["II_b2b"] * 2 + c["BR_b2b"] + c["SR_b2b"]},
            {"اسم المنتج": "مباني B2B", "عدد البكجات": c["HH_b2b"]},
            {"اسم المنتج": "درع B2B", "عدد البكجات": c["protect_shield_b2b"]},
            {"اسم المنتج": "ppf_q&e&p B2B", "عدد البكجات": ppf_b2b['ppf_e'] + ppf_b2b['ppf_q'] + ppf_b2b['ppf_p']},
            {"اسم المنتج": "ppf_f B2B", "عدد البكجات": ppf_b2b['ppf_f']},
        ]
        tables.append({
            "title": "عدد البكجات B2B",
            "columns": ["اسم المنتج", "عدد البكجات"],
            "rows": pkg_b2b_rows,
        })

        # ════════════════════════════════════════════════════════
        # 16. بكجات الدرع مع العازل
        # ════════════════════════════════════════════════════════
        opacity_protected = c["pdz"]
        pro_only = c["protect_shield"] - c["pdz"]
        tables.append({
            "title": "بكجات الدرع والعازل",
            "columns": ["بكجات التظليل مع الدرع", "بكجات الدرع بدون تظليل"],
            "rows": [{"بكجات التظليل مع الدرع": opacity_protected, "بكجات الدرع بدون تظليل": pro_only}],
        })

        # ════════════════════════════════════════════════════════
        # 17. داش كام
        # ════════════════════════════════════════════════════════
        tables.append({
            "title": "داش كام",
            "columns": ["داش كام"],
            "rows": [{"داش كام": c["dashCam"]}],
        })

        # ════════════════════════════════════════════════════════
        # 18. فحمات (Brakes)
        # ════════════════════════════════════════════════════════
        tables.append({
            "title": "فحمات",
            "columns": ["فحمات امامية", "فحمات خلفية", "المجموع"],
            "rows": [{
                "فحمات امامية": c["brake_front"],
                "فحمات خلفية": c["brake_rear"],
                "المجموع": c["brake_front"] + c["brake_rear"],
            }],
        })

        # ════════════════════════════════════════════════════════
        # 19. منتجات خاصة (Special products summary)
        # ════════════════════════════════════════════════════════
        special_rows = [
            {"المنتج": "CARD", "الكمية": c["CARD"]},
            {"المنتج": "SPRO3", "الكمية": c["SPRO3"]},
            {"المنتج": "MIC3", "الكمية": c["mic3"]},
            {"المنتج": "CLNo3", "الكمية": c["CLNo3"]},
        ]
        tables.append({
            "title": "منتجات خاصة",
            "columns": ["المنتج", "الكمية"],
            "rows": special_rows,
        })

        return {
            "tables": tables,
            "summary": {
                "total_orders": total_orders,
                "total_tint": total_b + total_s,
            },
        }

    # ── From Excel file ──────────────────────────────────────

    def compute_tables(self, df: pd.DataFrame) -> Dict[str, Any]:
        # Try skus_json column first (preferred, used by Salla export)
        json_col = None
        names_col = None

        if 'skus_json' in df.columns:
            json_col = 'skus_json'
        if 'اسماء المنتجات مع SKU' in df.columns:
            names_col = 'اسماء المنتجات مع SKU'

        # Determine which column and parser to use
        if json_col:
            sku_col = json_col
            use_json = True
        elif names_col:
            sku_col = names_col
            # Auto-detect if the names column contains JSON
            sample = df[sku_col].dropna().iloc[0] if len(df[sku_col].dropna()) > 0 else ""
            use_json = isinstance(sample, str) and sample.strip().startswith("[")
        else:
            # Fallback: search for any column with sku or منتج
            for col in df.columns:
                if "sku" in col.lower() or "منتج" in col:
                    sku_col = col
                    sample = df[sku_col].dropna().iloc[0] if len(df[sku_col].dropna()) > 0 else ""
                    use_json = isinstance(sample, str) and sample.strip().startswith("[")
                    break
            else:
                raise KeyError("الملف لا يحتوي على عمود المنتجات/SKU المطلوب")

        all_pairs = []
        for cell in df[sku_col]:
            parsed = self._parse_salla_cell(cell) if use_json else self._parse_names_column(cell)
            all_pairs.extend(parsed)

        counters = self._count_skus(all_pairs)
        return self._build_result_tables(counters, len(df))

    # ── From DB items ────────────────────────────────────────

    def compute_from_sku_list(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        parsed = [(self._split_sku(item["sku"]), item["quantity"]) for item in items]
        counters = self._count_skus(parsed)
        return self._build_result_tables(counters, 0)

    # ── Export ───────────────────────────────────────────────

    def export_excel(self, tables: Dict[str, Any], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Best Shield Report"
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        value_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        current_row = 1
        for table in tables.get("tables", []):
            # Title row
            title_cell = ws.cell(row=current_row, column=1, value=table["title"])
            title_cell.font = Font(bold=True, size=13)
            current_row += 1

            # Header row
            for col_idx, col_name in enumerate(table["columns"], 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Data rows
            for row_data in table["rows"]:
                for col_idx, col_name in enumerate(table["columns"], 1):
                    val = row_data.get(col_name, "")
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = value_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                current_row += 1

            current_row += 2  # Gap between tables

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        return output_path
